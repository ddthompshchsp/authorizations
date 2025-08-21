

import io
from datetime import datetime
from pathlib import Path
import re

import pandas as pd
import streamlit as st
import pytz
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="HCHSP — Disability Authorizations Formatter", layout="wide")

# ----------------------------
# Header (UI ONLY)
# ----------------------------
logo_path = Path("header_logo.png")
hdr_l, hdr_c, hdr_r = st.columns([1, 2, 1])
with hdr_c:
    if logo_path.exists():
        st.image(str(logo_path), width=320)
    st.markdown(
        """
        <h1 style='text-align:center; margin: 8px 0 4px;'>Hidalgo County Head Start — Disability Authorizations</h1>
        <p style='text-align:center; font-size:16px; margin-top:0;'>
        Upload the 10432 Quick Report export and download the cleaned, formatted workbook.
        </p>
        """,
        unsafe_allow_html=True,
    )

st.divider()

# ----------------------------
# File Upload
# ----------------------------
up = st.file_uploader("Upload *10432*.xlsx", type=["xlsx"], key="qf")

# ----------------------------
# Helpers
# ----------------------------
BLUE = "4472C4"  # header blue
WHITE = "FFFFFF"
GREEN = "008000"  # dates font
RED = "C00000"    # red X for missing dates

THIN = Side(style="thin", color="000000")
MED  = Side(style="medium", color="000000")


def _detect_header_row(df: pd.DataFrame) -> int:
    """Find row index that contains the header labels. Look for 'Participant PID'."""
    first_col = df.iloc[:, 0].astype(str).str.strip().str.lower()
    mask = first_col.str.contains("participant pid", na=False)
    if mask.any():
        return mask.idxmax()
    counts = df.apply(lambda r: r.astype(str).str.contains("ST:").sum(), axis=1)
    return int(counts.idxmax())


def _rename_columns(cols):
    mapping = {}
    for c in cols:
        c_str = str(c)
        c_norm = c_str.strip()
        if re.search(r"participant pid", c_norm, re.I):
            mapping[c] = "PID"
        elif re.search(r"participant first name", c_norm, re.I):
            mapping[c] = "First Name"
        elif re.search(r"participant last name", c_norm, re.I):
            mapping[c] = "Last Name"
        elif re.search(r"center name", c_norm, re.I):
            mapping[c] = "Center"
        elif re.search(r"class name", c_norm, re.I):
            mapping[c] = "Class"
        elif re.search(r"authorization.*date", c_norm, re.I):
            mapping[c] = "Authorization Date"
        else:
            mapping[c] = c_norm
    return [mapping[c] for c in cols]


def _autosize_columns(ws):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            val = cell.value
            length = 0 if val is None else len(str(val))
            if length > max_len:
                max_len = length
        ws.column_dimensions[letter].width = min(max_len + 2, 45)


def build_output_workbook(df: pd.DataFrame, title_text: str) -> bytes:
    """Build in-memory XLSX with styling and centered title (no logo)."""
    if "Authorization Date" in df.columns:
        dt = pd.to_datetime(df["Authorization Date"], errors="coerce")
        df["Authorization Date"] = dt.dt.strftime("%m/%d/%Y").fillna("")

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Authorizations", startrow=1)
        wb = writer.book
        ws = writer.sheets["Authorizations"]

        # Freeze panes below header row
        ws.freeze_panes = "A3"

        # Header style (row 2)
        header_row = 2
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.font = Font(bold=True, color=WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # AutoFilter
        ws.auto_filter.ref = ws.dimensions

        # Centered title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = title_text
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Borders
        max_row, max_col = ws.max_row, ws.max_column
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                left   = MED if c == 1 else THIN
                right  = MED if c == max_col else THIN
                top    = MED if r == 1 else THIN
                bottom = MED if r == max_row else THIN
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

        # Authorization Date column style
        if "Authorization Date" in df.columns:
            date_col_idx = list(df.columns).index("Authorization Date") + 1
            for r in range(3, ws.max_row + 1):
                cell = ws.cell(row=r, column=date_col_idx)
                val = (cell.value or "").strip() if isinstance(cell.value, str) else cell.value
                if val in (None, ""):
                    cell.value = "✗"
                    cell.font = Font(bold=True, color=RED)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.font = Font(color=GREEN)

        _autosize_columns(ws)

    return output.getvalue()


# ----------------------------
# Main
# ----------------------------
if up:
    safe_name = getattr(up, "name", "") or ""
    if "10432" not in safe_name:
        st.error("Please upload the correct file: filename must include **10432**.")
        st.stop()

    raw = pd.read_excel(up, header=None)
    hdr_row = _detect_header_row(raw)
    df = pd.read_excel(up, header=hdr_row)

    df = df.dropna(how="all")
    df.columns = _rename_columns(df.columns)

    desired_cols = ["PID", "First Name", "Last Name", "Center", "Class", "Authorization Date"]
    existing_cols = [c for c in desired_cols if c in df.columns]
    df = df[existing_cols]

    # Central time timestamp
    tz = pytz.timezone("America/Chicago")
    now_str = datetime.now(tz).strftime("%m/%d/%Y %I:%M %p")
    fixed_title = f"25-26 Authorizations — Exported {now_str}"

    xlsx_bytes = build_output_workbook(df, fixed_title)

    st.success("File processed. Click to download your formatted workbook.")
    st.download_button(
        label="Download Disability Authorizations (.xlsx)",
        data=xlsx_bytes,
        file_name=f"HCHSP_DisabilityAuthorizations_{datetime.now(tz).strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("Upload the **10432** Quick Report export (.xlsx) to begin.")

