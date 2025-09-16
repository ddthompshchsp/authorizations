import io
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import pytz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="HCHSP — Disability Authorizations (Services Style)", layout="wide")

# UI header (logo shown in Streamlit only; NOT embedded in Excel)
logo_path = Path("header_logo.png")
hdr_l, hdr_c, hdr_r = st.columns([1, 2, 1])
with hdr_c:
    if logo_path.exists():
        st.image(str(logo_path), width=320)
    st.markdown(
        """
        <h1 style='text-align:center; margin: 8px 0 4px;'>Hidalgo County Head Start — Disability Authorizations</h1>
        <p style='text-align:center; font-size:16px; margin-top:0;'>
        Upload the <b>10415 Authorization</b> export and download the cleaned, formatted workbook.
        </p>
        """,
        unsafe_allow_html=True,
    )
st.divider()

up = st.file_uploader("Upload *10415*.xlsx", type=["xlsx"], key="qf")

BLUE  = "1F4E78"
WHITE = "FFFFFF"
GRID  = "D9D9D9"
RED   = "C00000"
GREEN = "008000"
BLACK = "000000"

THIN = Side(style="thin",   color=BLACK)
MED  = Side(style="medium", color=BLACK)

def _autosize(ws, header_row: int):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        hv = ws.cell(row=header_row, column=col).value
        if hv is not None:
            max_len = len(str(hv))
        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 3, 48)

def _build(df: pd.DataFrame) -> bytes:
    # Normalize dates to strings
    if "Authorization Date" in df.columns:
        dt = pd.to_datetime(df["Authorization Date"], errors="coerce")
        df["Authorization Date"] = dt.dt.strftime("%m/%d/%Y").fillna("")

    wb = Workbook()
    ws = wb.active
    ws.title = "Authorizations"

    header_row = 4
    data_row0  = header_row + 1
    total_cols = max(1, df.shape[1])

    # Title (row 2) and subtitle (row 3)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)

    tcell = ws.cell(row=2, column=1, value="Hidalgo County Head Start Program")
    tcell.font = Font(bold=True, size=14)
    tcell.alignment = Alignment(horizontal="center", vertical="center")

    tz = pytz.timezone("America/Chicago")
    now_str = datetime.now(tz).strftime("%m/%d/%y %I:%M %p CT")
    scell = ws.cell(row=3, column=1, value=f"Disability Authorizations — 2025–2026 as of ({now_str})")
    scell.alignment = Alignment(horizontal="center", vertical="center")

    # Header row (row 4)
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=header_row, column=j, value=col)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.font = Font(bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows (start row 5) with banding and date styling
    for i, row in enumerate(df.itertuples(index=False), start=data_row0):
        is_band = (i - data_row0) % 2 == 1
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            if is_band:
                c.fill = PatternFill("solid", fgColor=GRID)
            if df.columns[j - 1] == "Authorization Date":
                if (val is None) or (str(val).strip() == ""):
                    c.value = "✗"
                    c.font = Font(bold=True, color=RED)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.font = Font(color=GREEN)

    max_row = ws.max_row
    max_col = ws.max_column

    # Thin grid
    for r in range(header_row, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Medium outline
    for c in range(1, max_col + 1):
        ws.cell(row=header_row, column=c).border = Border(
            left=MED if c == 1 else THIN,
            right=MED if c == max_col else THIN,
            top=MED,
            bottom=THIN,
        )
        ws.cell(row=max_row, column=c).border = Border(
            left=MED if c == 1 else THIN,
            right=MED if c == max_col else THIN,
            top=THIN,
            bottom=MED,
        )
    for r in range(header_row, max_row + 1):
        ws.cell(row=r, column=1).border       = Border(left=MED, right=THIN, top=THIN, bottom=THIN)
        ws.cell(row=r, column=max_col).border = Border(left=THIN, right=MED, top=THIN, bottom=THIN)

    # Freeze below header and put filter EXACTLY on the header row → sorting works
    ws.freeze_panes = f"A{data_row0}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"

    _autosize(ws, header_row)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def _detect_header_row(raw_df: pd.DataFrame) -> int:
    first_col = raw_df.iloc[:, 0].astype(str).str.strip().str.lower()
    if first_col.str.contains(r"authorization:\s*regarding my child", na=False).any():
        return first_col.str.contains(r"authorization:\s*regarding my child", na=False).idxmax()
    if first_col.str.contains("participant pid", na=False).any():
        return first_col.str.contains("participant pid", na=False).idxmax()
    return raw_df.notna().sum(axis=1).idxmax()

def _rename_columns(cols):
    mapping = {}
    for c in cols:
        s = str(c).strip()
        if re.search(r"authorization:\s*regarding my child", s, re.I):
            mapping[c] = "Child Name"
        elif re.search(r"authorization:\s*date", s, re.I):
            mapping[c] = "Authorization Date"
        elif re.search(r"IEP/IFSP\s*Dis:Identified", s, re.I):
            mapping[c] = "Disability Identified"
        elif re.search(r"primary\s*disability", s, re.I):
            mapping[c] = "Primary Disability"
        elif re.search(r"\bcenter name\b|\bcenter\b", s, re.I):
            mapping[c] = "Center"
        elif re.search(r"\bclass name\b|\bclass\b", s, re.I):
            mapping[c] = "Class"
        elif re.search(r"\bparticipant pid\b|\bpid\b", s, re.I):
            mapping[c] = "PID"
        elif re.search(r"\bfirst name\b", s, re.I):
            mapping[c] = "First Name"
        elif re.search(r"\blast name\b", s, re.I):
            mapping[c] = "Last Name"
        else:
            mapping[c] = s
    return [mapping[c] for c in cols]

def _split_child_name(df: pd.DataFrame) -> pd.DataFrame:
    if "Child Name" in df.columns:
        def split_name(full):
            if pd.isna(full):
                return pd.Series({"First Name": "", "Last Name": ""})
            parts = str(full).split()
            if len(parts) == 1:
                return pd.Series({"First Name": parts[0], "Last Name": ""})
            return pd.Series({"First Name": " ".join(parts[:-1]), "Last Name": parts[-1]})
        name_split = df["Child Name"].apply(split_name)
        for col in ["First Name", "Last Name"]:
            if col not in df.columns:
                df[col] = name_split[col]
    return df

# ----------------------------
# Main
# ----------------------------
if up:
    safe_name = getattr(up, "name", "") or ""
    if "10415" not in safe_name:
        st.error("Please upload the correct file: filename must include **10415**.")
        st.stop()

    raw = pd.read_excel(up, header=None)
    hdr_row = _detect_header_row(raw)

    df = pd.read_excel(up, header=hdr_row).dropna(how="all")
    df.columns = _rename_columns(df.columns)
    df = _split_child_name(df)

    preferred = ["PID", "First Name", "Last Name", "Center", "Class",
                 "Authorization Date", "Disability Identified", "Primary Disability"]
    existing = [c for c in preferred if c in df.columns]
    df = df[existing]

    xlsx = _build(df)

    st.success("File processed successfully. Click below to download.")
    st.download_button(
        "⬇️ Download Disability Authorizations (.xlsx)",
        data=xlsx,
        file_name=f"HCHSP_DisabilityAuthorizations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("Upload the **10415** export (.xlsx) to begin.")
